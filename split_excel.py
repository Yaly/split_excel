#!/usr/bin/python
# -*- coding: utf-8 -*-
# @Author: YALY
# @Date:   2017-01-13 15:11:17
# @Last Modified by:   yaly
# @Last Modified time: 2017-02-21 11:39:14
# Version 0.1

'''
Split the xlsx file.
The version 0.1 only support 2 columns file, you can custom this.
The number of rows per file is 100 by default, you can custom this too.
'''

from openpyxl import Workbook, load_workbook
import argparse


# define argumnts
parser = argparse.ArgumentParser()
parser.add_argument('-c', dest='file_num', default=100, help='specifiy the number of rows per file')
parser.add_argument('-f', dest='file_name', required = True, help='specify the Excel file')

args = parser.parse_args()


# Load xlsx file
wb1 = load_workbook(args.file_name)

# read worksheet
ws1 = wb1.active

# count work sheet rows
ws1_rows = len(tuple(ws1.rows))
ws1_columns = len(tuple(ws1.columns))


# set the number of rows per file
sep_rows = args.file_num


def insert_data(file_num, columns, rows=100):
	for row in range(1, rows+1):
		for column in range(1, columns+1):
			ws.cell(row = row, column=column).value = ws1.cell(row = row + file_num * 100, column = column).value

if ws1_rows % sep_rows == 0:
	file_nums = ws1_rows / sep_rows
else:
	file_nums =  ws1_rows / sep_rows + 1

for file_num in range(0, file_nums):
	wb = Workbook()
	ws = wb.active
	insert_data(file_num,ws1_columns, sep_rows)
	filename =  'wb_' + str(file_num) + '.xlsx'
	wb.save(filename=filename)